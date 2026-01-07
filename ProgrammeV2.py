#!/usr/bin/env python3
import re
import csv
from collections import Counter, defaultdict
from pathlib import Path
from datetime import datetime

import tkinter as tk
from tkinter import filedialog, messagebox

import webbrowser
import markdown  # pip install markdown
import base64
import io
import matplotlib.pyplot as plt  # pip install matplotlib

# ===== Excel avec graphes (openpyxl) =====
try:
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------
# CSS bento box pour le rapport HTML
# ---------------------------------------------------------

css_moderne = """
* {
  box-sizing: border-box;
}

html, body {
  margin: 0;
  padding: 0;
}

body {
  font-family: system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
  background: #f3f4f6;
  color: #111827;
}

/* Barre de haut de page */
header {
  padding: 20px 24px;
  border-bottom: 1px solid #e5e7eb;
  background: #f9fafb;
}

header h1 {
  margin: 0;
  font-size: 1.7rem;
  font-weight: 600;
}

header p {
  margin: 4px 0 0 0;
  font-size: 0.9rem;
  color: #6b7280;
}

/* Conteneur principal */
main {
  max-width: 1100px;
  margin: 0 auto;
  padding: 20px 16px 32px 16px;
  display: flex;
  flex-direction: column;
  gap: 16px;
}

/* Cartes bento */
.bento-grid {
  display: flex;
  flex-direction: column;
  gap: 16px;
}

/* Carte générique */
.card {
  background: #ffffff;
  border-radius: 14px;
  padding: 18px 18px 20px 18px;
  box-shadow: 0 10px 25px rgba(15, 23, 42, 0.08);
}

/* Section Markdown dans une carte */
.card-md {
  background: #ffffff;
  border-radius: 14px;
  padding: 18px 18px 20px 18px;
  box-shadow: 0 10px 25px rgba(15, 23, 42, 0.06);
}

/* Titres */
h1, h2, h3 {
  font-weight: 600;
  color: #111827;
}

h2 {
  font-size: 1.25rem;
  margin-top: 0;
  margin-bottom: 10px;
}

h3 {
  font-size: 1.05rem;
  margin-top: 16px;
  margin-bottom: 6px;
}

/* Texte */
p {
  line-height: 1.7;
  font-size: 0.95rem;
  margin: 6px 0;
}

/* Tableaux */
table {
  border-collapse: collapse;
  width: 100%;
  margin: 14px 0;
  font-size: 0.9rem;
}

th, td {
  border: 1px solid #e5e7eb;
  padding: 6px 8px;
  text-align: left;
}

th {
  background: #f9fafb;
  font-weight: 600;
}

/* Graphes bento sur une ligne */
.graphs-row {
  display: flex;
  flex-wrap: wrap;
  gap: 14px;
}

.graph-card {
  background: #f9fafb;
  border-radius: 12px;
  padding: 10px 10px 12px 10px;
  flex: 1 1 30%;
  min-width: 260px;
  box-shadow: 0 4px 12px rgba(15, 23, 42, 0.04);
}

.graph-card-title {
  font-size: 0.85rem;
  color: #4b5563;
  margin-bottom: 6px;
}

.graph-card img {
  max-width: 100%;
  height: auto;
  border-radius: 8px;
}

/* Listes et code */
ul, ol {
  padding-left: 20px;
}

li {
  margin: 3px 0;
}

code {
  font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Courier New", monospace;
  font-size: 0.85rem;
  background: #f3f4f6;
  padding: 1px 4px;
  border-radius: 4px;
}

/* Bas de page */
footer {
  background: #f9fafb;
  border-top: 1px solid #e5e7eb;
  padding: 14px 16px;
  margin-top: 18px;
  font-size: 0.85rem;
  color: #6b7280;
  text-align: center;
}

/* Responsive */
@media (max-width: 800px) {
  main {
    padding: 16px 10px 24px 10px;
  }
  .graphs-row {
    flex-direction: column;
  }
  .graph-card {
    min-width: 100%;
  }
}
"""

# ---------------------------------------------------------
# Thème Tkinter bento-like (fond gris + cartes)
# ---------------------------------------------------------

def apply_tk_theme(root):
    """Thème type bento : fond gris, cartes blanches arrondies, boutons plats."""
    bg_main = "#f3f4f6"
    card_bg = "#ffffff"
    border = "#e5e7eb"
    text_main = "#111827"
    text_sub = "#6b7280"

    root.configure(bg=bg_main)

    default_font = ("Segoe UI", 10)
    root.option_add("*Font", default_font)
    root.option_add("*Background", card_bg)
    root.option_add("*Label.Background", card_bg)
    root.option_add("*Label.Foreground", text_main)
    root.option_add("*Frame.Background", card_bg)
    root.option_add("*Button.Background", card_bg)
    root.option_add("*Button.Foreground", text_main)
    root.option_add("*Button.BorderWidth", 1)
    root.option_add("*Button.Relief", "solid")
    root.option_add("*Entry.Background", "#f9fafb")
    root.option_add("*Entry.Relief", "solid")
    root.option_add("*Entry.BorderWidth", 1)

    return {
        "bg_main": bg_main,
        "card_bg": card_bg,
        "border": border,
        "text_main": text_main,
        "text_sub": text_sub,
    }

# ---------------------------------------------------------
# Configuration langues (FR / EN)
# ---------------------------------------------------------

current_lang = "fr"
lang_var = None

TEXTS = {
    "fr": {
        "app_title": "Analyse tcpdump - SAE 1.05",
        "menu_file": "Fichier",
        "menu_open": "Ouvrir...",
        "menu_save_md": "Enregistrer le rapport (.md)...",
        "menu_save_csv": "Enregistrer les CSV...",
        "menu_save_xlsx": "Enregistrer l'Excel (.xlsx)...",
        "menu_quit": "Quitter",
        "btn_choose": "Choisir un fichier tcpdump",
        "btn_save_md": "Enregistrer le rapport (.md)",
        "btn_save_csv": "Enregistrer les CSV",
        "btn_save_xlsx": "Enregistrer l'Excel (.xlsx)",
        "btn_quit": "Quitter",
        "btn_view_html": "Voir le rapport",
        "thresholds_frame": "Seuils de détection (modifiable)",
        "th_scan_ports": "Ports distincts (scan) ≥",
        "th_dos_abs": "Paquets vers une IP (DoS) ≥",
        "th_dos_pct": "Pourcentage trafic (DoS) ≥",
        "th_noisy_dests": "Destinations distinctes (noisy) ≥",
        "th_noisy_pkts": "Paquets par source (noisy) ≥",
        "th_syn_abs": "Paquets SYN (SYN flood) ≥",
        "th_syn_ratio": "Ratio SYN flood (%)",
        "no_file": "Aucun fichier sélectionné",
        "dlg_open_title": "Sélectionner un fichier tcpdump",
        "dlg_open_types": "Fichiers texte",
        "warn_no_packets_title": "Analyse",
        "warn_no_packets_msg": "Aucun paquet n'a été parsé dans ce fichier.",
        "err_thresholds_title": "Erreur",
        "err_thresholds_msg": "Les seuils doivent être des nombres valides.",
        "info_done_title": "Analyse terminée",
        "info_done_msg": "Analyse terminée.\nPaquets : {packets}\nAlertes détectées : {alerts}",
        "warn_no_md_title": "Sauvegarde",
        "warn_no_md_msg": "Aucun rapport Markdown à sauvegarder.",
        "dlg_save_md_title": "Enregistrer le rapport Markdown",
        "dlg_save_md_type": "Fichiers Markdown",
        "info_md_saved": "Rapport sauvegardé dans :\n{path}",
        "warn_no_csv_title": "Sauvegarde",
        "warn_no_csv_msg": "Aucun CSV à sauvegarder.",
        "dlg_save_csv_title": "Choisir un dossier pour les CSV",
        "dlg_save_csv_type": "Dossier",
        "info_csv_saved": "CSV sauvegardés dans :\n{path}",
        "warn_no_xlsx_title": "Excel",
        "warn_no_xlsx_msg": "openpyxl n'est pas installé.\nInstalle-le avec : pip install openpyxl",
        "dlg_save_xlsx_title": "Enregistrer l'Excel avec graphes",
        "dlg_save_xlsx_type": "Fichiers Excel",
        "info_xlsx_saved": "Excel avec graphes sauvegardé dans :\n{path}",
        "err_save_title": "Erreur",
        "err_save_msg": "Erreur lors de l'enregistrement :\n{error}",
        "summary_none": (
            "Aucune activité suspecte n'a été clairement mise en évidence avec "
            "les seuils actuels. Le trafic observé semble globalement normal.\n"
        ),
        "summary_intro": "Voici un résumé des principaux éléments détectés dans la capture :\n",
        "summary_scan_title": "- Suspicion de **scan de ports** :\n",
        "summary_scan_line": (
            "  - {src} semble tester un grand nombre de ports sur {dst}.\n"
        ),
        "summary_dos_title": "- Suspicion de **déni de service (DoS)** :\n",
        "summary_dos_line": (
            "  - Une grosse partie du trafic est dirigée vers {dst}, "
            "ce qui peut indiquer une tentative de saturation.\n"
        ),
        "summary_syn_title": "- Suspicion de **SYN flood** :\n",
        "summary_syn_line": (
            "  - Le nombre de paquets SYN vers {dst} est très élevé "
            "par rapport au reste du trafic.\n"
        ),
        "summary_noisy_title": "- Présence de **sources très bavardes** :\n",
        "summary_noisy_line": (
            "  - {src} émet un volume important de paquets ou contacte "
            "beaucoup de destinations différentes.\n"
        ),
        "summary_context": (
            "\nCes observations doivent être replacées dans le contexte du réseau :\n"
            "- il peut s'agir d'attaques réelles,\n"
            "- ou d'applications légitimes très actives (sauvegardes, mises à jour, etc.).\n"
        ),
        "md_title": "# Rapport d'analyse du trafic réseau\n\n",
        "md_summary_title": "## Résumé général\n\n",
        "md_info_title": "## Informations générales sur la capture\n\n",
        "lang_label": "Langue :",
        "lang_fr": "FR",
        "lang_en": "EN",
    },
    "en": {
        "app_title": "tcpdump analysis - SAE 1.05",
        "menu_file": "File",
        "menu_open": "Open...",
        "menu_save_md": "Save report (.md)...",
        "menu_save_csv": "Save CSVs...",
        "menu_save_xlsx": "Save Excel (.xlsx)...",
        "menu_quit": "Quit",
        "btn_choose": "Choose tcpdump file",
        "btn_save_md": "Save report (.md)",
        "btn_save_csv": "Save CSVs",
        "btn_save_xlsx": "Save Excel (.xlsx)",
        "btn_quit": "Quit",
        "btn_view_html": "View report",
        "thresholds_frame": "Detection thresholds (editable)",
        "th_scan_ports": "Distinct ports (scan) ≥",
        "th_dos_abs": "Packets to one IP (DoS) ≥",
        "th_dos_pct": "Traffic percentage (DoS) ≥",
        "th_noisy_dests": "Distinct destinations (noisy) ≥",
        "th_noisy_pkts": "Packets per source (noisy) ≥",
        "th_syn_abs": "SYN packets (SYN flood) ≥",
        "th_syn_ratio": "SYN flood ratio (%)",
        "no_file": "No file selected",
        "dlg_open_title": "Select tcpdump file",
        "dlg_open_types": "Text files",
        "warn_no_packets_title": "Analysis",
        "warn_no_packets_msg": "No packets were parsed from this file.",
        "err_thresholds_title": "Error",
        "err_thresholds_msg": "Thresholds must be valid numbers.",
        "info_done_title": "Analysis finished",
        "info_done_msg": "Analysis finished.\nPackets: {packets}\nAlerts detected: {alerts}",
        "warn_no_md_title": "Save",
        "warn_no_md_msg": "No Markdown report to save.",
        "dlg_save_md_title": "Save Markdown report",
        "dlg_save_md_type": "Markdown files",
        "info_md_saved": "Report saved to:\n{path}",
        "warn_no_csv_title": "Save",
        "warn_no_csv_msg": "No CSV to save.",
        "dlg_save_csv_title": "Choose a folder for CSVs",
        "dlg_save_csv_type": "Folder",
        "info_csv_saved": "CSVs saved to:\n{path}",
        "warn_no_xlsx_title": "Excel",
        "warn_no_xlsx_msg": "openpyxl is not installed.\nInstall it with: pip install openpyxl",
        "dlg_save_xlsx_title": "Save Excel with charts",
        "dlg_save_xlsx_type": "Excel files",
        "info_xlsx_saved": "Excel with charts saved to:\n{path}",
        "err_save_title": "Error",
        "err_save_msg": "Error while saving:\n{error}",
        "summary_none": (
            "No clearly suspicious activity was found with the current thresholds. "
            "The observed traffic looks mostly normal.\n"
        ),
        "summary_intro": "Here is a summary of the main elements detected in the capture:\n",
        "summary_scan_title": "- Suspected **port scan**:\n",
        "summary_scan_line": (
            "  - {src} seems to be testing many ports on {dst}.\n"
        ),
        "summary_dos_title": "- Suspected **denial of service (DoS)**:\n",
        "summary_dos_line": (
            "  - A large part of the traffic is directed to {dst}, "
            "which may indicate an attempt to overload it.\n"
        ),
        "summary_syn_title": "- Suspected **SYN flood**:\n",
        "summary_syn_line": (
            "  - The number of SYN packets to {dst} is very high "
            "compared to the rest of the traffic.\n"
        ),
        "summary_noisy_title": "- Presence of **very talkative sources**:\n",
        "summary_noisy_line": (
            "  - {src} sends a large number of packets or contacts many different destinations.\n"
        ),
        "summary_context": (
            "\nThese observations must be interpreted in the context of the network:\n"
            "- they may correspond to real attacks,\n"
            "- or to legitimate but very active applications (backups, updates, etc.).\n"
        ),
        "md_title": "# Network traffic analysis report\n\n",
        "md_summary_title": "## General summary\n\n",
        "md_info_title": "## General information about the capture\n\n",
        "lang_label": "Language:",
        "lang_fr": "FR",
        "lang_en": "EN",
    },
}

def switch_language():
    global current_lang
    current_lang = lang_var.get()

# ---------------------------------------------------------
# Parsing tcpdump
# ---------------------------------------------------------

LINE_RE = re.compile(
    r'^(?P<time>\d{2}:\d{2}:\d{2}\.\d+)\s+IP6?\s+'
    r'(?P<src>[^ >]+)\s*>\s*(?P<dst>[^:]+):\s*'
    r'(?:Flags\s*\[(?P<flags>[^\]]*)\].*?)?'
    r'(?:(?:length|len)\s+(?P<length>\d+))?'
)

HOST_PORT_RE = re.compile(r'^(?P<host>.+)\.(?P<port>[^.]+)$')

def split_host_port(field: str):
    m = HOST_PORT_RE.match(field)
    if not m:
        return field, None
    return m.group('host'), m.group('port')

def parse_tcpdump_file(path):
    packets = []
    first_ts = None
    last_ts = None

    with open(path, encoding='utf-8', errors='ignore') as f:
        for line in f:
            line = line.strip()
            if not line or not line[0].isdigit():
                continue

            proto = "OTHER"
            if " ICMP " in line:
                proto = "ICMP"
            elif " UDP " in line:
                proto = "UDP"
            elif " Flags [" in line or " tcp " in line.lower():
                proto = "TCP"

            m = LINE_RE.match(line)
            if not m:
                continue

            gd = m.groupdict()
            time_str = gd['time']
            src_raw = gd['src']
            dst_raw = gd['dst']
            flags = gd.get('flags') or ''
            length = gd.get('length')
            length = int(length) if length is not None else 0

            try:
                ts = datetime.strptime(time_str, "%H:%M:%S.%f")
            except ValueError:
                ts = None

            if ts:
                if first_ts is None or ts < first_ts:
                    first_ts = ts
                if last_ts is None or ts > last_ts:
                    last_ts = ts

            src_host, src_port = split_host_port(src_raw)
            dst_host, dst_port = split_host_port(dst_raw)

            packets.append({
                'time': time_str,
                'src_host': src_host,
                'src_port': src_port,
                'dst_host': dst_host,
                'dst_port': dst_port,
                'flags': flags,
                'length': length,
                'proto': proto,
            })

    return packets, first_ts, last_ts

# ---------------------------------------------------------
# Stats & détection
# ---------------------------------------------------------

def is_syn_flag(flags: str) -> bool:
    if not flags:
        return False
    return 'S' in flags

def compute_stats(packets):
    by_src = Counter()
    by_dst = Counter()
    by_dport = Counter()
    by_flow = Counter()
    dests_per_src = defaultdict(set)
    total_bytes = 0
    flags_counter = Counter()
    syn_per_dst = Counter()
    syn_total = 0
    by_proto = Counter()

    for p in packets:
        sh = p['src_host']
        dh = p['dst_host']
        dp = p['dst_port']
        fl = p['flags']
        proto = p.get('proto', 'OTHER')

        by_src[sh] += 1
        by_dst[dh] += 1
        by_proto[proto] += 1
        if dp:
            by_dport[dp] += 1
        by_flow[(sh, dh, dp)] += 1
        dests_per_src[sh].add(dh)
        total_bytes += p['length']

        if fl:
            flags_counter[fl] += 1
            if is_syn_flag(fl):
                syn_per_dst[dh] += 1
                syn_total += 1

    flows_per_src = Counter({s: len(dests) for s, dests in dests_per_src.items()})
    return by_src, by_dst, by_dport, by_flow, flows_per_src, total_bytes, flags_counter, syn_per_dst, syn_total, by_proto

def detect_port_scans(by_flow, threshold_ports=20):
    ports_per_pair = defaultdict(set)
    for (sh, dh, dp), _ in by_flow.items():
        if dp is None:
            continue
        ports_per_pair[(sh, dh)].add(dp)

    alerts = []
    for (sh, dh), ports in ports_per_pair.items():
        if len(ports) >= threshold_ports:
            alerts.append({
                'type': 'PORT_SCAN',
                'src': sh,
                'dst': dh,
                'unique_dst_ports': len(ports),
            })
    return alerts

def detect_dos(by_dst, abs_threshold=1000, pct_threshold=0.3):
    alerts = []
    total = sum(by_dst.values())

    for dst, count in by_dst.items():
        ratio = count / total if total else 0
        if count >= abs_threshold or ratio >= pct_threshold:
            alerts.append({
                'type': 'POSSIBLE_DOS',
                'dst': dst,
                'packets': count,
                'ratio': ratio,
            })
    return alerts

def detect_noisy_sources(by_src, flows_per_src, dest_threshold=50, pkt_threshold=80):
    alerts = []
    for src in by_src.keys():
        n_dests = flows_per_src.get(src, 0)
        pkts = by_src[src]
        if n_dests >= dest_threshold or pkts >= pkt_threshold:
            alerts.append({
                'type': 'NOISY_SOURCE',
                'src': src,
                'distinct_dests': n_dests,
                'packets': pkts,
            })
    return alerts

def detect_syn_flood(syn_per_dst, by_dst, syn_abs=500, syn_ratio=0.5):
    alerts = []
    for dst, syn_count in syn_per_dst.items():
        total_to_dst = by_dst.get(dst, 0)
        if total_to_dst == 0:
            continue
        ratio = syn_count / total_to_dst
        if syn_count >= syn_abs and ratio >= syn_ratio:
            alerts.append({
                'type': 'POSSIBLE_SYN_FLOOD',
                'dst': dst,
                'syn_packets': syn_count,
                'total_packets_to_dst': total_to_dst,
                'syn_ratio': ratio,
            })
    return alerts

# ---------------------------------------------------------
# CSV
# ---------------------------------------------------------

def save_csv():
    t = TEXTS[current_lang]
    if not last_packets:
        messagebox.showwarning(t["warn_no_csv_title"], t["warn_no_csv_msg"])
        return

    folder = filedialog.askdirectory(title=t["dlg_save_csv_title"])
    if not folder:
        return

    try:
        folder_path = Path(folder)

        raw_path = folder_path / "packets_raw.csv"
        with open(raw_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(last_packets[0].keys()))
            writer.writeheader()
            writer.writerows(last_packets)

        (by_src, by_dst, by_dport, by_flow,
         flows_per_src, total_bytes, flags_counter,
         syn_per_dst, syn_total, by_proto) = last_stats

        src_path = folder_path / "stats_sources.csv"
        with open(src_path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["src_host", "packets"])
            for host, count in by_src.most_common():
                w.writerow([host, count])

        dst_path = folder_path / "stats_destinations.csv"
        with open(dst_path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["dst_host", "packets"])
            for host, count in by_dst.most_common():
                w.writerow([host, count])

        port_path = folder_path / "stats_ports.csv"
        with open(port_path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["dst_port", "packets"])
            for port, count in by_dport.most_common():
                w.writerow([port, count])

        syn_path = folder_path / "stats_syn_per_dst.csv"
        with open(syn_path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["dst_host", "syn_packets"])
            for dst, syn_count in syn_per_dst.most_common():
                w.writerow([dst, syn_count])

        proto_path = folder_path / "stats_protocoles.csv"
        with open(proto_path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["protocole", "packets"])
            for proto, count in by_proto.most_common():
                w.writerow([proto, count])

        messagebox.showinfo(t["warn_no_csv_title"], t["info_csv_saved"].format(path=folder_path))
    except Exception as e:
        messagebox.showerror(t["err_save_title"], t["err_save_msg"].format(error=e))

# ---------------------------------------------------------
# Excel avec graphes
# ---------------------------------------------------------

def save_excel_with_charts():
    t = TEXTS[current_lang]

    if not HAS_OPENPYXL:
        messagebox.showwarning(t["warn_no_xlsx_title"], t["warn_no_xlsx_msg"])
        return

    if not last_packets or not last_stats:
        messagebox.showwarning("Excel", "Aucune analyse en mémoire.\nLance d'abord une analyse de fichier.")
        return

    filename = filedialog.asksaveasfilename(
        title=t["dlg_save_xlsx_title"],
        defaultextension=".xlsx",
        filetypes=((t["dlg_save_xlsx_type"], "*.xlsx"), ("Tous les fichiers", "*.*"))
    )
    if not filename:
        return

    try:
        (by_src, by_dst, by_dport, by_flow,
         flows_per_src, total_bytes, flags_counter,
         syn_per_dst, syn_total, by_proto) = last_stats

        wb = Workbook()
        wb.remove(wb.active)

        ws_raw = wb.create_sheet("Raw")
        headers = list(last_packets[0].keys())
        ws_raw.append(headers)
        for p in last_packets:
            ws_raw.append([p.get(h, "") for h in headers])

        ws_src = wb.create_sheet("Sources")
        ws_src.append(["src_host", "packets"])
        for host, count in by_src.most_common():
            ws_src.append([host, count])

        pie_src = PieChart()
        pie_src.title = "Trafic par source"
        labels = Reference(ws_src, min_col=1, min_row=2, max_row=ws_src.max_row)
        data = Reference(ws_src, min_col=2, min_row=1, max_row=ws_src.max_row)
        pie_src.add_data(data, titles_from_data=True)
        pie_src.set_categories(labels)
        ws_src.add_chart(pie_src, "E2")

        ws_dst = wb.create_sheet("Destinations")
        ws_dst.append(["dst_host", "packets"])
        for host, count in by_dst.most_common():
            ws_dst.append([host, count])

        pie_dst = PieChart()
        pie_dst.title = "Trafic par destination"
        labels = Reference(ws_dst, min_col=1, min_row=2, max_row=ws_dst.max_row)
        data = Reference(ws_dst, min_col=2, min_row=1, max_row=ws_dst.max_row)
        pie_dst.add_data(data, titles_from_data=True)
        pie_dst.set_categories(labels)
        ws_dst.add_chart(pie_dst, "E2")

        ws_port = wb.create_sheet("Ports")
        ws_port.append(["dst_port", "packets"])
        for port, count in by_dport.most_common():
            ws_port.append([str(port), count])

        bar_port = BarChart()
        bar_port.title = "Ports de destination les plus utilisés"
        labels = Reference(ws_port, min_col=1, min_row=2, max_row=ws_port.max_row)
        data = Reference(ws_port, min_col=2, min_row=1, max_row=ws_port.max_row)
        bar_port.add_data(data, titles_from_data=True)
        bar_port.set_categories(labels)
        bar_port.y_axis.title = "Paquets"
        bar_port.x_axis.title = "Port"
        ws_port.add_chart(bar_port, "E2")

        ws_syn = wb.create_sheet("SYN")
        ws_syn.append(["dst_host", "syn_packets"])
        for dst, syn_count in syn_per_dst.most_common():
            ws_syn.append([dst, syn_count])

        bar_syn = BarChart()
        bar_syn.title = "Paquets SYN par destination"
        labels = Reference(ws_syn, min_col=1, min_row=2, max_row=ws_syn.max_row)
        data = Reference(ws_syn, min_col=2, min_row=1, max_row=ws_syn.max_row)
        bar_syn.add_data(data, titles_from_data=True)
        bar_syn.set_categories(labels)
        bar_syn.y_axis.title = "Paquets SYN"
        ws_syn.add_chart(bar_syn, "E2")

        ws_sum = wb.create_sheet("Résumé")
        ws_sum.append(["Indicateur", "Valeur"])
        ws_sum.append(["Paquets analysés", len(last_packets)])
        ws_sum.append(["Volume total (octets)", total_bytes])
        ws_sum.append(["Sources distinctes", len(by_src)])
        ws_sum.append(["Destinations distinctes", len(by_dst)])
        ws_sum.append(["Ports distincts", len(by_dport)])
        ws_sum.append(["Protocoles distincts", len(by_proto)])
        ws_sum.append(["Paquets SYN totaux", syn_total])
        ws_sum.append(["Alertes détectées", len(last_alerts) if last_alerts else 0])

        wb.save(filename)
        messagebox.showinfo(t["warn_no_xlsx_title"], t["info_xlsx_saved"].format(path=filename))
    except Exception as e:
        messagebox.showerror(t["err_save_title"], t["err_save_msg"].format(error=e))

# ---------------------------------------------------------
# Graphes matplotlib -> HTML
# ---------------------------------------------------------

def fig_to_base64(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode("ascii")
    plt.close(fig)
    return b64

def build_chart_images(by_src, by_dst, by_dport):
    images = {}

    if by_dst:
        labels = []
        sizes = []
        for dst, count in by_dst.most_common(5):
            labels.append(str(dst))
            sizes.append(count)
        other = sum(c for _, c in list(by_dst.items())[5:])
        if other > 0:
            labels.append("Autres")
            sizes.append(other)
        fig1, ax1 = plt.subplots(figsize=(4, 4))
        ax1.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
        ax1.set_title("Répartition du trafic par destination")
        images["traffic_by_dst"] = fig_to_base64(fig1)

    if by_src:
        src_labels = [str(h) for h, _ in by_src.most_common(8)]
        src_values = [c for _, c in by_src.most_common(8)]
        fig2, ax2 = plt.subplots(figsize=(5, 3))
        ax2.bar(range(len(src_labels)), src_values)
        ax2.set_xticks(range(len(src_labels)))
        ax2.set_xticklabels(src_labels, rotation=45, ha="right", fontsize=8)
        ax2.set_ylabel("Paquets")
        ax2.set_title("Principales sources de trafic")
        images["top_sources"] = fig_to_base64(fig2)

    if by_dport:
        port_labels = [str(p) for p, _ in by_dport.most_common(8)]
        port_values = [c for _, c in by_dport.most_common(8)]
        fig3, ax3 = plt.subplots(figsize=(5, 3))
        ax3.bar(range(len(port_labels)), port_values)
        ax3.set_xticks(range(len(port_labels)))
        ax3.set_xticklabels(port_labels, rotation=45, ha="right", fontsize=8)
        ax3.set_ylabel("Paquets")
        ax3.set_title("Ports de destination les plus utilisés")
        images["top_ports"] = fig_to_base64(fig3)

    return images

# ---------------------------------------------------------
# Résumé & rapport Markdown
# ---------------------------------------------------------

def host_label(ip_or_name):
    if any(c.isalpha() for c in ip_or_name):
        return ip_or_name
    if current_lang == "fr":
        return f"hôte {ip_or_name}"
    return f"host {ip_or_name}"

def build_summary(alerts):
    t = TEXTS[current_lang]
    if not alerts:
        return t["summary_none"]

    scan_entries = []
    dos_entries = []
    syn_entries = []
    noisy_entries = []

    for a in alerts:
        if a["type"] == "PORT_SCAN":
            src_lab = host_label(a["src"])
            dst_lab = host_label(a["dst"])
            line = t["summary_scan_line"].format(src=src_lab, dst=dst_lab)
            if line not in scan_entries:
                scan_entries.append(line)
        elif a["type"] == "POSSIBLE_DOS":
            dst_lab = host_label(a["dst"])
            line = t["summary_dos_line"].format(dst=dst_lab)
            if line not in dos_entries:
                dos_entries.append(line)
        elif a["type"] == "POSSIBLE_SYN_FLOOD":
            dst_lab = host_label(a["dst"])
            line = t["summary_syn_line"].format(dst=dst_lab)
            if line not in syn_entries:
                syn_entries.append(line)
        elif a["type"] == "NOISY_SOURCE":
            src_lab = host_label(a["src"])
            line = t["summary_noisy_line"].format(src=src_lab)
            if line not in noisy_entries:
                noisy_entries.append(line)

    lines = []
    lines.append(t["summary_intro"])
    lines.append("\n")

    if scan_entries:
        lines.append(t["summary_scan_title"])
        for l in scan_entries:
            lines.append(l)
        lines.append("\n")

    if dos_entries:
        lines.append(t["summary_dos_title"])
        for l in dos_entries:
            lines.append(l)
        lines.append("\n")

    if syn_entries:
        lines.append(t["summary_syn_title"])
        for l in syn_entries:
            lines.append(l)
        lines.append("\n")

    if noisy_entries:
        lines.append(t["summary_noisy_title"])
        for l in noisy_entries:
            lines.append(l)
        lines.append("\n")

    lines.append(t["summary_context"])
    return "".join(lines)

def build_markdown_report(packets, by_src, by_dst, by_dport, by_flow,
                          alerts, first_ts, last_ts, total_bytes, flags_counter,
                          syn_per_dst, syn_total, by_proto):
    t = TEXTS[current_lang]
    total_packets = len(packets)

    if first_ts and last_ts and last_ts > first_ts:
        duration_sec = (last_ts - first_ts).total_seconds()
    else:
        duration_sec = 0.0

    pps = total_packets / duration_sec if duration_sec > 0 else 0.0

    lines = []
    lines.append(t["md_title"])
    lines.append(t["md_summary_title"])
    lines.append(build_summary(alerts) + "\n")

    lines.append(t["md_info_title"])
    lines.append("| Élément | Valeur |\n")
    lines.append("|--------|--------|\n")
    lines.append(f"| Paquets analysés | {total_packets} |\n")
    lines.append(f"| Volume total | {total_bytes} octets |\n")
    if first_ts and last_ts:
        lines.append(f"| Premier paquet | `{first_ts.time()}` |\n")
        lines.append(f"| Dernier paquet | `{last_ts.time()}` |\n")
        lines.append(f"| Durée de la capture | {duration_sec:.3f} s |\n")
    else:
        lines.append("| Horodatages | Non disponibles |\n")
    lines.append(f"| Débit moyen | {pps:.3f} paquets/s |\n")
    lines.append(f"| Paquets SYN totaux | {syn_total} |\n\n")

    lines.append("## Répartition par protocole\n\n")
    if not by_proto:
        lines.append("Aucun protocole identifié.\n\n")
    else:
        lines.append("| Protocole | Paquets |\n")
        lines.append("|-----------|---------|\n")
        for proto, count in by_proto.most_common():
            lines.append(f"| {proto} | {count} |\n")
        lines.append("\n")

    lines.append("## Comportement des connexions TCP\n\n")
    if not flags_counter:
        lines.append("Aucun flag TCP identifié.\n\n")
    else:
        lines.append("| Flags | Nombre de paquets |\n")
        lines.append("|-------|--------------------|\n")
        for fl, c in flags_counter.most_common():
            lines.append(f"| `{fl}` | {c} |\n")
        lines.append("\n")

    lines.append("## Répartition des paquets SYN par destination\n\n")
    if not syn_per_dst:
        lines.append("Aucun paquet SYN distinct.\n\n")
    else:
        lines.append("| Destination | Paquets SYN | Paquets totaux | Ratio SYN |\n")
        lines.append("|-------------|------------:|---------------:|-----------:|\n")
        for dst, syn_count in syn_per_dst.most_common():
            total_to_dst = by_dst.get(dst, 0)
            ratio = syn_count / total_to_dst if total_to_dst else 0
            lines.append(f"| {dst} | {syn_count} | {total_to_dst} | {ratio:.2f} |\n")
        lines.append("\n")

    lines.append("## Hôtes émetteurs principaux\n\n")
    lines.append("| Source | Paquets |\n")
    lines.append("|--------|---------|\n")
    for host, count in by_src.most_common(10):
        lines.append(f"| {host} | {count} |\n")
    lines.append("\n")

    lines.append("## Hôtes destinataires principaux\n\n")
    lines.append("| Destination | Paquets |\n")
    lines.append("|-------------|---------|\n")
    for host, count in by_dst.most_common(10):
        lines.append(f"| {host} | {count} |\n")
    lines.append("\n")

    lines.append("## Ports de destination les plus utilisés\n\n")
    lines.append("| Port | Paquets |\n")
    lines.append("|------|---------|\n")
    for port, count in by_dport.most_common(10):
        lines.append(f"| {port} | {count} |\n")
    lines.append("\n")

    lines.append("## Alertes détectées\n\n")
    if not alerts:
        lines.append("Aucune alerte avec les seuils actuels.\n")
    else:
        lines.append("| Type | Détail |\n")
        lines.append("|------|--------|\n")
        for a in alerts:
            if a["type"] == "PORT_SCAN":
                detail = f"{host_label(a['src'])} teste {a['unique_dst_ports']} ports sur {host_label(a['dst'])}"
            elif a["type"] == "POSSIBLE_DOS":
                pct = round(a["ratio"] * 100, 1)
                detail = f"DoS sur {host_label(a['dst'])}: {a['packets']} paquets ({pct}%)"
            elif a["type"] == "POSSIBLE_SYN_FLOOD":
                pct = round(a["syn_ratio"] * 100, 1)
                detail = f"SYN flood sur {host_label(a['dst'])}: {a['syn_packets']} SYN ({pct}%)"
            elif a["type"] == "NOISY_SOURCE":
                detail = f"{host_label(a['src'])} → {a['distinct_dests']} dests / {a['packets']} pkts"
            else:
                detail = str(a)
            lines.append(f"| {a['type']} | {detail} |\n")

    lines.append("\n## Activités suspectes principales\n\n")
    if not alerts:
        lines.append("Aucune activité clairement suspecte avec les seuils actuels.\n")
    else:
        top_descriptions = []
        for a in alerts:
            if a["type"] == "PORT_SCAN":
                desc = f"Scan de ports depuis {host_label(a['src'])} vers {host_label(a['dst'])}."
            elif a["type"] == "POSSIBLE_DOS":
                pct = round(a["ratio"] * 100, 1)
                desc = f"Possible DoS vers {host_label(a['dst'])} avec {a['packets']} paquets ({pct} % du trafic)."
            elif a["type"] == "POSSIBLE_SYN_FLOOD":
                pct = round(a["syn_ratio"] * 100, 1)
                desc = f"Possible SYN flood vers {host_label(a['dst'])} avec {a['syn_packets']} SYN ({pct} % des paquets vers cette cible)."
            elif a["type"] == "NOISY_SOURCE":
                desc = f"Source très bavarde : {host_label(a['src'])} vers {a['distinct_dests']} destinations différentes."
            else:
                desc = str(a)
            if desc not in top_descriptions:
                top_descriptions.append(desc)
            if len(top_descriptions) >= 2:
                break

        if top_descriptions:
            for d in top_descriptions:
                lines.append(f"- {d}\n")
        else:
            lines.append("Les alertes détectées ne permettent pas d'isoler deux activités dominantes.\n")

    return "".join(lines)

# ---------------------------------------------------------
# GUI / état global
# ---------------------------------------------------------

current_markdown = ""
last_packets = None
last_first_ts = None
last_last_ts = None
last_stats = None
last_alerts = None
chart_images_global = {}

DEFAULT_SCAN_PORTS = 20
DEFAULT_DOS_ABS = 1000
DEFAULT_DOS_PCT = 30
DEFAULT_NOISY_DESTS = 50
DEFAULT_NOISY_PKTS = 80
DEFAULT_SYN_ABS = 500
DEFAULT_SYN_RATIO = 50

entry_scan_ports = None
entry_dos_abs = None
entry_dos_pct = None
entry_noisy_dests = None
entry_noisy_pkts = None
entry_syn_abs = None
entry_syn_ratio = None
label_chemin = None
text_md = None

def analyser_fichier():
    global current_markdown
    global last_packets, last_first_ts, last_last_ts, last_stats, last_alerts
    global chart_images_global

    t = TEXTS[current_lang]

    chemin_fichier = filedialog.askopenfilename(
        title=t["dlg_open_title"],
        filetypes=((t["dlg_open_types"], "*.txt"), ("All files", "*.*"))
    )
    if not chemin_fichier:
        return

    label_chemin.config(text=chemin_fichier)

    in_path = Path(chemin_fichier)
    packets, first_ts, last_ts = parse_tcpdump_file(in_path)

    if not packets:
        messagebox.showwarning(t["warn_no_packets_title"], t["warn_no_packets_msg"])
        return

    try:
        th_scan = int(entry_scan_ports.get() or DEFAULT_SCAN_PORTS)
        th_dos_abs = int(entry_dos_abs.get() or DEFAULT_DOS_ABS)
        th_dos_pct = float(entry_dos_pct.get() or DEFAULT_DOS_PCT) / 100.0
        th_noisy_dests = int(entry_noisy_dests.get() or DEFAULT_NOISY_DESTS)
        th_noisy_pkts = int(entry_noisy_pkts.get() or DEFAULT_NOISY_PKTS)
        th_syn_abs = int(entry_syn_abs.get() or DEFAULT_SYN_ABS)
        th_syn_ratio = float(entry_syn_ratio.get() or DEFAULT_SYN_RATIO) / 100.0
    except ValueError:
        messagebox.showerror(t["err_thresholds_title"], t["err_thresholds_msg"])
        return

    stats = compute_stats(packets)
    (by_src, by_dst, by_dport, by_flow,
     flows_per_src, total_bytes, flags_counter,
     syn_per_dst, syn_total, by_proto) = stats

    scans = detect_port_scans(by_flow, threshold_ports=th_scan)
    dos = detect_dos(by_dst, abs_threshold=th_dos_abs, pct_threshold=th_dos_pct)
    noisy = detect_noisy_sources(by_src, flows_per_src,
                                 dest_threshold=th_noisy_dests,
                                 pkt_threshold=th_noisy_pkts)
    syn_floods = detect_syn_flood(syn_per_dst, by_dst,
                                  syn_abs=th_syn_abs,
                                  syn_ratio=th_syn_ratio)

    alerts = scans + dos + noisy + syn_floods

    chart_images_global = build_chart_images(by_src, by_dst, by_dport)

    last_packets = packets
    last_first_ts = first_ts
    last_last_ts = last_ts
    last_stats = stats
    last_alerts = alerts

    current_markdown = build_markdown_report(
        packets, by_src, by_dst, by_dport, by_flow,
        alerts, first_ts, last_ts, total_bytes, flags_counter,
        syn_per_dst, syn_total, by_proto
    )

    if first_ts and last_ts and last_ts > first_ts:
        duration_sec = (last_ts - first_ts).total_seconds()
    else:
        duration_sec = 0.0
    pps = len(packets) / duration_sec if duration_sec > 0 else 0.0

    text_md.config(state="normal")
    text_md.delete("1.0", tk.END)
    resume = (
        f"Fichier : {chemin_fichier}\n"
        f"Paquets analysés : {len(packets)}\n"
        f"Volume total : {total_bytes} octets\n"
        f"Durée de la capture : {duration_sec:.3f} s\n"
        f"Débit moyen : {pps:.3f} paquets/s\n"
        f"Paquets SYN : {syn_total}\n"
        f"Alertes détectées : {len(alerts)}\n\n"
        "Utilise les boutons ci-dessous pour exporter le rapport, les CSV ou l'Excel."
    )
    text_md.insert("1.0", resume)
    text_md.config(state="disabled")

    messagebox.showinfo(
        t["info_done_title"],
        t["info_done_msg"].format(packets=len(packets), alerts=len(alerts))
    )

def save_markdown():
    t = TEXTS[current_lang]
    if not current_markdown:
        messagebox.showwarning(t["warn_no_md_title"], t["warn_no_md_msg"])
        return

    filename = filedialog.asksaveasfilename(
        title=t["dlg_save_md_title"],
        defaultextension=".md",
        filetypes=((t["dlg_save_md_type"], "*.md"), ("All files", "*.*"))
    )
    if not filename:
        return

    try:
        with open(filename, "w", encoding="utf-8") as f:
            f.write(current_markdown)
        messagebox.showinfo(t["warn_no_md_title"], t["info_md_saved"].format(path=filename))
    except Exception as e:
        messagebox.showerror(t["err_save_title"], t["err_save_msg"].format(error=e))

def open_report_in_browser():
    t = TEXTS[current_lang]
    if not current_markdown:
        messagebox.showwarning(t["warn_no_md_title"], t["warn_no_md_msg"])
        return

    html_body = markdown.markdown(current_markdown, extensions=['tables'])

    graphs_html = ""
    if chart_images_global:
        graphs_html += """
    <section class="card">
      <h2>Graphes récapitulatifs</h2>
      <div class="graphs-row">
"""
        mapping = [
            ("traffic_by_dst", "Répartition du trafic par destination"),
            ("top_sources", "Principales sources de trafic"),
            ("top_ports", "Ports de destination les plus utilisés"),
        ]
        for key, title in mapping:
            img = chart_images_global.get(key)
            if img:
                graphs_html += f"""
        <div class="graph-card">
          <div class="graph-card-title">{title}</div>
          <img src="data:image/png;base64,{img}" alt="{title}">
        </div>
"""
        graphs_html += """
      </div>
    </section>
"""

    html = f"""<!DOCTYPE html>
<html lang="{current_lang}">
<head>
<meta charset="utf-8">
<title>Rapport SAE 1.05</title>
<style>
{css_moderne}
</style>
</head>
<body>
<header>
  <h1>SAE 1.05 – Analyse de trafic réseau</h1>
  <p>BUT Réseaux & Télécoms – Année 1 – Étudiant : Prénom NOM</p>
</header>
<main>
<div class="bento-grid">
{graphs_html}
<section class="card-md">
{html_body}
</section>
</div>
</main>
<footer>
  <p>&copy; 2026 – Analyse de trafic réseau – SAE 1.05</p>
</footer>
</body>
</html>
"""

    import tempfile
    with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
        f.write(html)
        temp_file = f.name

    webbrowser.open('file://' + temp_file)

# ---------------------------------------------------------
# Interface Tkinter (bento)
# ---------------------------------------------------------

def create_gui():
    global lang_var, text_md
    global entry_scan_ports, entry_dos_abs, entry_dos_pct
    global entry_noisy_dests, entry_noisy_pkts
    global entry_syn_abs, entry_syn_ratio
    global label_chemin

    root = tk.Tk()
    t = TEXTS[current_lang]
    root.title(t["app_title"])
    root.geometry("900x650")

    colors = apply_tk_theme(root)

    # fond gris derrière les cartes
    root.configure(bg=colors["bg_main"])

    menu_bar = tk.Menu(root)
    root.config(menu=menu_bar)

    file_menu = tk.Menu(menu_bar, tearoff=0)
    menu_bar.add_cascade(label=t["menu_file"], menu=file_menu)
    file_menu.add_command(label=t["menu_open"], command=analyser_fichier)
    file_menu.add_separator()
    file_menu.add_command(label=t["menu_save_md"], command=save_markdown)
    file_menu.add_command(label=t["menu_save_csv"], command=save_csv)
    file_menu.add_command(label=t["menu_save_xlsx"], command=save_excel_with_charts)
    file_menu.add_separator()
    file_menu.add_command(label=t["menu_quit"], command=root.quit)

    outer = tk.Frame(root, bg=colors["bg_main"])
    outer.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)

    # carte principale bento
    card = tk.Frame(
        outer,
        bg=colors["card_bg"],
        bd=0,
        highlightbackground=colors["border"],
        highlightthickness=1
    )
    card.pack(fill=tk.BOTH, expand=True)
    card.pack_propagate(False)

    inner = tk.Frame(card, bg=colors["card_bg"])
    inner.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)

    # langue
    lang_frame = tk.Frame(inner, bg=colors["card_bg"])
    lang_frame.pack(fill=tk.X, pady=(0, 8))
    tk.Label(lang_frame, text=t["lang_label"], fg=colors["text_sub"], bg=colors["card_bg"]).pack(side=tk.LEFT)
    lang_var = tk.StringVar(value=current_lang)
    tk.Radiobutton(
        lang_frame, text=t["lang_fr"], variable=lang_var, value="fr",
        command=switch_language, bg=colors["card_bg"]
    ).pack(side=tk.LEFT, padx=(8, 0))
    tk.Radiobutton(
        lang_frame, text=t["lang_en"], variable=lang_var, value="en",
        command=switch_language, bg=colors["card_bg"]
    ).pack(side=tk.LEFT, padx=(4, 0))

    # carte fichier
    frame_file = tk.Frame(inner, bg=colors["card_bg"])
    frame_file.pack(fill=tk.X, pady=(4, 10))

    tk.Label(frame_file, text="Fichier à analyser", fg=colors["text_main"], bg=colors["card_bg"]).pack(
        anchor="w", pady=(0, 4)
    )

    file_row = tk.Frame(frame_file, bg=colors["card_bg"])
    file_row.pack(fill=tk.X)

    btn_choose = tk.Button(
        file_row, text=t["btn_choose"], command=analyser_fichier,
        relief="solid", bd=1, bg=colors["card_bg"], activebackground="#e5e7eb"
    )
    btn_choose.pack(side=tk.LEFT)

    label_chemin = tk.Label(
        file_row, text=t["no_file"], fg=colors["text_sub"], bg=colors["card_bg"], anchor="w"
    )
    label_chemin.pack(side=tk.LEFT, padx=(10, 0), fill=tk.X, expand=True)

    # carte seuils
    frame_th = tk.Frame(inner, bg=colors["card_bg"])
    frame_th.pack(fill=tk.X, pady=(0, 10))

    tk.Label(frame_th, text=t["thresholds_frame"], fg=colors["text_main"], bg=colors["card_bg"]).grid(
        row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 8)
    )

    row = 1
    tk.Label(frame_th, text=t["th_scan_ports"], bg=colors["card_bg"]).grid(row=row, column=0, sticky=tk.W, pady=3)
    entry_scan_ports = tk.Entry(frame_th, width=10)
    entry_scan_ports.insert(0, str(DEFAULT_SCAN_PORTS))
    entry_scan_ports.grid(row=row, column=1, sticky=tk.W, padx=(10, 0))

    row += 1
    tk.Label(frame_th, text=t["th_dos_abs"], bg=colors["card_bg"]).grid(row=row, column=0, sticky=tk.W, pady=3)
    entry_dos_abs = tk.Entry(frame_th, width=10)
    entry_dos_abs.insert(0, str(DEFAULT_DOS_ABS))
    entry_dos_abs.grid(row=row, column=1, sticky=tk.W, padx=(10, 0))

    row += 1
    tk.Label(frame_th, text=t["th_dos_pct"], bg=colors["card_bg"]).grid(row=row, column=0, sticky=tk.W, pady=3)
    entry_dos_pct = tk.Entry(frame_th, width=10)
    entry_dos_pct.insert(0, str(DEFAULT_DOS_PCT))
    entry_dos_pct.grid(row=row, column=1, sticky=tk.W, padx=(10, 0))

    row += 1
    tk.Label(frame_th, text=t["th_noisy_dests"], bg=colors["card_bg"]).grid(row=row, column=0, sticky=tk.W, pady=3)
    entry_noisy_dests = tk.Entry(frame_th, width=10)
    entry_noisy_dests.insert(0, str(DEFAULT_NOISY_DESTS))
    entry_noisy_dests.grid(row=row, column=1, sticky=tk.W, padx=(10, 0))

    row += 1
    tk.Label(frame_th, text=t["th_noisy_pkts"], bg=colors["card_bg"]).grid(row=row, column=0, sticky=tk.W, pady=3)
    entry_noisy_pkts = tk.Entry(frame_th, width=10)
    entry_noisy_pkts.insert(0, str(DEFAULT_NOISY_PKTS))
    entry_noisy_pkts.grid(row=row, column=1, sticky=tk.W, padx=(10, 0))

    row += 1
    tk.Label(frame_th, text=t["th_syn_abs"], bg=colors["card_bg"]).grid(row=row, column=0, sticky=tk.W, pady=3)
    entry_syn_abs = tk.Entry(frame_th, width=10)
    entry_syn_abs.insert(0, str(DEFAULT_SYN_ABS))
    entry_syn_abs.grid(row=row, column=1, sticky=tk.W, padx=(10, 0))

    row += 1
    tk.Label(frame_th, text=t["th_syn_ratio"], bg=colors["card_bg"]).grid(row=row, column=0, sticky=tk.W, pady=3)
    entry_syn_ratio = tk.Entry(frame_th, width=10)
    entry_syn_ratio.insert(0, str(DEFAULT_SYN_RATIO))
    entry_syn_ratio.grid(row=row, column=1, sticky=tk.W, padx=(10, 0))

    # carte résumé
    frame_md = tk.Frame(inner, bg=colors["card_bg"])
    frame_md.pack(fill=tk.BOTH, expand=True, pady=(8, 0))

    tk.Label(frame_md, text="Résumé de l'analyse", fg=colors["text_main"], bg=colors["card_bg"]).pack(
        anchor="w", pady=(0, 4)
    )

    text_container = tk.Frame(frame_md, bg=colors["card_bg"])
    text_container.pack(fill=tk.BOTH, expand=True)

    text_md_scroll = tk.Scrollbar(text_container)
    text_md_scroll.pack(side=tk.RIGHT, fill=tk.Y)

    text_md = tk.Text(
        text_container, height=10, wrap=tk.WORD, state="disabled",
        bg="#f9fafb", bd=1, relief="solid"
    )
    text_md.pack(fill=tk.BOTH, expand=True)
    text_md_scroll.config(command=text_md.yview)
    text_md.config(yscrollcommand=text_md_scroll.set)

    # boutons bento en bas
    frame_btn = tk.Frame(inner, bg=colors["card_bg"])
    frame_btn.pack(fill=tk.X, pady=(10, 0))

    def styled_button(parent, text, cmd):
        return tk.Button(
            parent, text=text, command=cmd,
            relief="solid", bd=1, bg=colors["card_bg"], activebackground="#e5e7eb"
        )

    styled_button(frame_btn, t["btn_view_html"], open_report_in_browser).pack(side=tk.LEFT)
    styled_button(frame_btn, t["btn_save_md"], save_markdown).pack(side=tk.LEFT, padx=(8, 0))
    styled_button(frame_btn, t["btn_save_csv"], save_csv).pack(side=tk.LEFT, padx=(8, 0))
    styled_button(frame_btn, t["btn_save_xlsx"], save_excel_with_charts).pack(side=tk.LEFT, padx=(8, 0))
    styled_button(frame_btn, t["btn_quit"], root.quit).pack(side=tk.RIGHT)

    root.mainloop()

if __name__ == "__main__":
    create_gui()
